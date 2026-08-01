from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.database import engine
from app.services.report_branding_service import (
    ReportBranding,
    get_report_branding,
)


# -------------------------------------------------------------------
# Fixed workbook colors
# -------------------------------------------------------------------

BLUE_COLOR = "2563EB"
AMBER_COLOR = "D97706"
RED_COLOR = "DC2626"

LIGHT_BACKGROUND = "F8FAFC"
BORDER_COLOR = "CBD5E1"
WHITE_COLOR = "FFFFFF"
TEXT_COLOR = "0F172A"
MUTED_TEXT_COLOR = "64748B"

GREEN_LIGHT = "DCFCE7"
AMBER_LIGHT = "FEF3C7"
RED_LIGHT = "FEE2E2"


# -------------------------------------------------------------------
# Database helpers
# -------------------------------------------------------------------

def _fetch_rows(query: str) -> List[Dict[str, Any]]:
    """Execute a read-only query and return rows as dictionaries."""

    with engine.connect() as connection:
        result = connection.execute(text(query))
        return [dict(row._mapping) for row in result]


def _fetch_production_data() -> List[Dict[str, Any]]:
    return _fetch_rows(
        """
        SELECT
            report_date,
            mine_name,
            ore_plan,
            ore_actual,
            waste_plan,
            waste_actual,
            created_at
        FROM public.production_daily
        ORDER BY report_date ASC
        """
    )


def _fetch_fleet_data() -> List[Dict[str, Any]]:
    return _fetch_rows(
        """
        SELECT
            report_date,
            mine_name,
            availability,
            utilization,
            created_at
        FROM public.fleet_daily
        ORDER BY report_date ASC
        """
    )


def _fetch_plant_data() -> List[Dict[str, Any]]:
    return _fetch_rows(
        """
        SELECT
            report_date,
            mine_name,
            throughput_plan,
            throughput_actual,
            recovery,
            created_at
        FROM public.plant_daily
        ORDER BY report_date ASC
        """
    )


def _fetch_safety_data() -> List[Dict[str, Any]]:
    return _fetch_rows(
        """
        SELECT
            report_date,
            mine_name,
            incidents,
            near_misses,
            critical_risks,
            safety_score,
            created_at
        FROM public.safety_daily
        ORDER BY report_date ASC
        """
    )


# -------------------------------------------------------------------
# Value helpers
# -------------------------------------------------------------------

def _number(value: Any) -> float:
    if value is None:
        return 0.0

    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _safe_ratio(numerator: Any, denominator: Any) -> Optional[float]:
    denominator_value = _number(denominator)

    if denominator_value == 0:
        return None

    return _number(numerator) / denominator_value


def _average(values: Iterable[Any]) -> Optional[float]:
    cleaned_values = [
        _number(value)
        for value in values
        if value is not None
    ]

    if not cleaned_values:
        return None

    return sum(cleaned_values) / len(cleaned_values)


def _latest_row(rows: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    return rows[-1] if rows else None


def _latest_report_date(
    datasets: Iterable[List[Dict[str, Any]]],
) -> Optional[date]:
    dates: List[date] = []

    for rows in datasets:
        for row in rows:
            value = row.get("report_date")

            if isinstance(value, datetime):
                dates.append(value.date())
            elif isinstance(value, date):
                dates.append(value)

    return max(dates) if dates else None


def _display_value(value: Optional[float], suffix: str = "") -> str:
    if value is None:
        return "No data"

    if suffix == "%":
        return f"{value:.1f}%"

    return f"{value:,.1f}{suffix}"


# -------------------------------------------------------------------
# Dynamic workbook styles
# -------------------------------------------------------------------

def _build_styles(branding: ReportBranding) -> Dict[str, Any]:
    primary_color = branding.primary_color_excel
    secondary_color = branding.secondary_color_excel

    thin_border = Border(
        left=Side(style="thin", color=BORDER_COLOR),
        right=Side(style="thin", color=BORDER_COLOR),
        top=Side(style="thin", color=BORDER_COLOR),
        bottom=Side(style="thin", color=BORDER_COLOR),
    )

    return {
        "primary_color": primary_color,
        "secondary_color": secondary_color,
        "thin_border": thin_border,
        "header_fill": PatternFill(
            fill_type="solid",
            fgColor=primary_color,
        ),
        "section_fill": PatternFill(
            fill_type="solid",
            fgColor=secondary_color,
        ),
        "light_fill": PatternFill(
            fill_type="solid",
            fgColor=LIGHT_BACKGROUND,
        ),
        "blue_fill": PatternFill(
            fill_type="solid",
            fgColor=BLUE_COLOR,
        ),
        "green_fill": PatternFill(
            fill_type="solid",
            fgColor=GREEN_LIGHT,
        ),
        "amber_fill": PatternFill(
            fill_type="solid",
            fgColor=AMBER_LIGHT,
        ),
        "red_fill": PatternFill(
            fill_type="solid",
            fgColor=RED_LIGHT,
        ),
    }


# -------------------------------------------------------------------
# Workbook formatting helpers
# -------------------------------------------------------------------

def _set_sheet_view(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"


def _set_column_widths(
    worksheet,
    widths: Dict[str, float],
) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _style_title(
    worksheet,
    title: str,
    subtitle: str,
    styles: Dict[str, Any],
    end_column: int,
) -> int:
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=end_column,
    )

    title_cell = worksheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(
        name="Arial",
        size=18,
        bold=True,
        color=WHITE_COLOR,
    )
    title_cell.fill = styles["header_fill"]
    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    title_cell.border = styles["thin_border"]
    worksheet.row_dimensions[1].height = 30

    worksheet.merge_cells(
        start_row=2,
        start_column=1,
        end_row=2,
        end_column=end_column,
    )

    subtitle_cell = worksheet.cell(
        row=2,
        column=1,
        value=subtitle,
    )
    subtitle_cell.font = Font(
        name="Arial",
        size=10,
        color=MUTED_TEXT_COLOR,
        italic=True,
    )
    subtitle_cell.fill = styles["light_fill"]
    subtitle_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    subtitle_cell.border = styles["thin_border"]
    worksheet.row_dimensions[2].height = 22

    return 4


def _style_section_heading(
    worksheet,
    row: int,
    title: str,
    styles: Dict[str, Any],
    end_column: int,
) -> None:
    worksheet.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=end_column,
    )

    cell = worksheet.cell(row=row, column=1, value=title)
    cell.font = Font(
        name="Arial",
        size=11,
        bold=True,
        color=WHITE_COLOR,
    )
    cell.fill = styles["section_fill"]
    cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    cell.border = styles["thin_border"]
    worksheet.row_dimensions[row].height = 22


def _style_table_header(
    worksheet,
    row: int,
    headers: List[str],
    styles: Dict[str, Any],
) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=row,
            column=column_index,
            value=header,
        )
        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=WHITE_COLOR,
        )
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = styles["thin_border"]

    worksheet.row_dimensions[row].height = 30


def _style_data_row(
    worksheet,
    row: int,
    column_count: int,
    styles: Dict[str, Any],
) -> None:
    for column_index in range(1, column_count + 1):
        cell = worksheet.cell(row=row, column=column_index)
        cell.font = Font(
            name="Arial",
            size=10,
            color=TEXT_COLOR,
        )
        cell.alignment = Alignment(
            vertical="center",
        )
        cell.border = styles["thin_border"]

        if row % 2 == 0:
            cell.fill = styles["light_fill"]


def _apply_date_format(cell) -> None:
    if isinstance(cell.value, (date, datetime)):
        cell.number_format = "yyyy-mm-dd"


def _apply_datetime_format(cell) -> None:
    if isinstance(cell.value, datetime):
        cell.number_format = "yyyy-mm-dd hh:mm"


def _apply_number_format(cell, decimals: int = 1) -> None:
    cell.number_format = f'#,##0.{"0" * decimals}'


def _apply_percentage_format(cell, decimals: int = 1) -> None:
    cell.number_format = f'0.{"0" * decimals}%'


def _add_autofilter(
    worksheet,
    header_row: int,
    last_row: int,
    last_column: int,
) -> None:
    if last_row <= header_row:
        return

    last_column_letter = get_column_letter(last_column)
    worksheet.auto_filter.ref = (
        f"A{header_row}:{last_column_letter}{last_row}"
    )


def _add_logo(
    worksheet,
    branding: ReportBranding,
    anchor: str = "F1",
    width: int = 130,
    height: int = 45,
) -> None:
    if not branding.logo_path:
        return

    logo_path = Path(branding.logo_path)

    if not logo_path.exists():
        return

    try:
        from openpyxl.drawing.image import Image as ExcelImage

        image = ExcelImage(str(logo_path))
        image.width = width
        image.height = height
        worksheet.add_image(image, anchor)
    except Exception:
        # Branding should never prevent report generation.
        return


# -------------------------------------------------------------------
# Executive summary worksheet
# -------------------------------------------------------------------

def _create_executive_summary_sheet(
    workbook: Workbook,
    branding: ReportBranding,
    styles: Dict[str, Any],
    production_rows: List[Dict[str, Any]],
    fleet_rows: List[Dict[str, Any]],
    plant_rows: List[Dict[str, Any]],
    safety_rows: List[Dict[str, Any]],
) -> None:
    worksheet = workbook.active
    worksheet.title = "Executive Summary"
    worksheet.sheet_view.showGridLines = False

    worksheet.merge_cells("A1:F2")
    title_cell = worksheet["A1"]
    title_cell.value = branding.company_name
    title_cell.font = Font(
        name="Arial",
        size=20,
        bold=True,
        color=WHITE_COLOR,
    )
    title_cell.fill = styles["header_fill"]
    title_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    title_cell.border = styles["thin_border"]

    worksheet.merge_cells("A3:F3")
    report_cell = worksheet["A3"]
    report_cell.value = "Executive Operations Intelligence Export"
    report_cell.font = Font(
        name="Arial",
        size=14,
        bold=True,
        color=TEXT_COLOR,
    )
    report_cell.fill = styles["light_fill"]
    report_cell.alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    report_cell.border = styles["thin_border"]

    _add_logo(
        worksheet,
        branding,
        anchor="E1",
        width=115,
        height=38,
    )

    latest_date = _latest_report_date(
        [
            production_rows,
            fleet_rows,
            plant_rows,
            safety_rows,
        ]
    )

    metadata = [
        ("Mine", branding.mine_name),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        (
            "Latest operational data",
            latest_date.strftime("%Y-%m-%d")
            if latest_date
            else "No data",
        ),
        ("Timezone", branding.timezone),
        ("Language", branding.language),
    ]

    start_row = 5

    for offset, (label, value) in enumerate(metadata):
        row = start_row + offset

        label_cell = worksheet.cell(row=row, column=1, value=label)
        label_cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=MUTED_TEXT_COLOR,
        )
        label_cell.fill = styles["light_fill"]
        label_cell.border = styles["thin_border"]

        worksheet.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=6,
        )

        value_cell = worksheet.cell(row=row, column=2, value=value)
        value_cell.font = Font(
            name="Arial",
            size=10,
            color=TEXT_COLOR,
        )
        value_cell.border = styles["thin_border"]
        value_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        for column in range(2, 7):
            worksheet.cell(row=row, column=column).border = (
                styles["thin_border"]
            )

    section_row = start_row + len(metadata) + 1
    _style_section_heading(
        worksheet,
        section_row,
        "Executive KPI Snapshot",
        styles,
        6,
    )

    latest_production = _latest_row(production_rows)
    latest_fleet = _latest_row(fleet_rows)
    latest_plant = _latest_row(plant_rows)
    latest_safety = _latest_row(safety_rows)

    ore_achievement = (
        _safe_ratio(
            latest_production.get("ore_actual"),
            latest_production.get("ore_plan"),
        )
        if latest_production
        else None
    )

    waste_achievement = (
        _safe_ratio(
            latest_production.get("waste_actual"),
            latest_production.get("waste_plan"),
        )
        if latest_production
        else None
    )

    plant_achievement = (
        _safe_ratio(
            latest_plant.get("throughput_actual"),
            latest_plant.get("throughput_plan"),
        )
        if latest_plant
        else None
    )

    kpis = [
        (
            "Ore plan achievement",
            _display_value(
                ore_achievement * 100
                if ore_achievement is not None
                else None,
                "%",
            ),
            ore_achievement,
        ),
        (
            "Waste plan achievement",
            _display_value(
                waste_achievement * 100
                if waste_achievement is not None
                else None,
                "%",
            ),
            waste_achievement,
        ),
        (
            "Fleet availability",
            _display_value(
                _number(latest_fleet.get("availability"))
                if latest_fleet
                else None,
                "%",
            ),
            (
                _number(latest_fleet.get("availability")) / 100
                if latest_fleet
                else None
            ),
        ),
        (
            "Fleet utilization",
            _display_value(
                _number(latest_fleet.get("utilization"))
                if latest_fleet
                else None,
                "%",
            ),
            (
                _number(latest_fleet.get("utilization")) / 100
                if latest_fleet
                else None
            ),
        ),
        (
            "Plant throughput achievement",
            _display_value(
                plant_achievement * 100
                if plant_achievement is not None
                else None,
                "%",
            ),
            plant_achievement,
        ),
        (
            "Safety score",
            _display_value(
                _number(latest_safety.get("safety_score"))
                if latest_safety
                else None,
                "%",
            ),
            (
                _number(latest_safety.get("safety_score")) / 100
                if latest_safety
                else None
            ),
        ),
    ]

    header_row = section_row + 1
    _style_table_header(
        worksheet,
        header_row,
        ["KPI", "Current Value", "Status"],
        styles,
    )

    for index, (kpi_name, display_value, ratio_value) in enumerate(
        kpis,
        start=header_row + 1,
    ):
        status = "No data"

        if ratio_value is not None:
            if ratio_value >= 1:
                status = "On or above target"
            elif ratio_value >= 0.9:
                status = "Watch"
            else:
                status = "Below target"

        worksheet.cell(row=index, column=1, value=kpi_name)
        worksheet.cell(row=index, column=2, value=display_value)
        worksheet.cell(row=index, column=3, value=status)

        worksheet.merge_cells(
            start_row=index,
            start_column=3,
            end_row=index,
            end_column=6,
        )

        _style_data_row(
            worksheet,
            index,
            6,
            styles,
        )

        status_cell = worksheet.cell(row=index, column=3)

        if status == "On or above target":
            status_cell.fill = styles["green_fill"]
        elif status == "Watch":
            status_cell.fill = styles["amber_fill"]
        elif status == "Below target":
            status_cell.fill = styles["red_fill"]

    data_row = header_row + len(kpis) + 2
    _style_section_heading(
        worksheet,
        data_row,
        "Data Coverage",
        styles,
        6,
    )

    coverage = [
        ("Production records", len(production_rows)),
        ("Fleet records", len(fleet_rows)),
        ("Plant records", len(plant_rows)),
        ("Safety records", len(safety_rows)),
    ]

    for offset, (label, count) in enumerate(coverage, start=1):
        row = data_row + offset
        worksheet.cell(row=row, column=1, value=label)
        worksheet.cell(row=row, column=2, value=count)
        worksheet.merge_cells(
            start_row=row,
            start_column=2,
            end_row=row,
            end_column=6,
        )
        _style_data_row(worksheet, row, 6, styles)

    _set_column_widths(
        worksheet,
        {
            "A": 31,
            "B": 19,
            "C": 20,
            "D": 16,
            "E": 16,
            "F": 16,
        },
    )

    worksheet.freeze_panes = "A5"
    worksheet.print_title_rows = "1:3"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins.left = 0.3
    worksheet.page_margins.right = 0.3
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5


# -------------------------------------------------------------------
# Operational worksheets
# -------------------------------------------------------------------

def _create_production_sheet(
    workbook: Workbook,
    branding: ReportBranding,
    styles: Dict[str, Any],
    rows: List[Dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet("Production")
    worksheet.sheet_view.showGridLines = False

    subtitle = (
        f"{branding.company_name} | {branding.mine_name} | "
        "Daily production performance"
    )
    header_row = _style_title(
        worksheet,
        "Production Performance",
        subtitle,
        styles,
        10,
    )

    headers = [
        "Report Date",
        "Mine",
        "Ore Plan",
        "Ore Actual",
        "Ore Variance",
        "Ore Achievement",
        "Waste Plan",
        "Waste Actual",
        "Waste Variance",
        "Created At",
    ]

    _style_table_header(
        worksheet,
        header_row,
        headers,
        styles,
    )

    for row_index, record in enumerate(rows, start=header_row + 1):
        ore_plan = _number(record.get("ore_plan"))
        ore_actual = _number(record.get("ore_actual"))
        waste_plan = _number(record.get("waste_plan"))
        waste_actual = _number(record.get("waste_actual"))

        values = [
            record.get("report_date"),
            record.get("mine_name") or branding.mine_name,
            ore_plan,
            ore_actual,
            ore_actual - ore_plan,
            _safe_ratio(ore_actual, ore_plan),
            waste_plan,
            waste_actual,
            waste_actual - waste_plan,
            record.get("created_at"),
        ]

        for column_index, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

        _style_data_row(
            worksheet,
            row_index,
            len(headers),
            styles,
        )

        _apply_date_format(worksheet.cell(row=row_index, column=1))

        for column_index in [3, 4, 5, 7, 8, 9]:
            _apply_number_format(
                worksheet.cell(
                    row=row_index,
                    column=column_index,
                )
            )

        _apply_percentage_format(
            worksheet.cell(row=row_index, column=6)
        )
        _apply_datetime_format(
            worksheet.cell(row=row_index, column=10)
        )

    last_row = header_row + len(rows)
    _add_autofilter(
        worksheet,
        header_row,
        last_row,
        len(headers),
    )

    if last_row > header_row:
        achievement_range = f"F{header_row + 1}:F{last_row}"
        worksheet.conditional_formatting.add(
            achievement_range,
            CellIsRule(
                operator="lessThan",
                formula=["0.90"],
                fill=styles["red_fill"],
            ),
        )
        worksheet.conditional_formatting.add(
            achievement_range,
            CellIsRule(
                operator="between",
                formula=["0.90", "0.999999"],
                fill=styles["amber_fill"],
            ),
        )
        worksheet.conditional_formatting.add(
            achievement_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["1.00"],
                fill=styles["green_fill"],
            ),
        )

    _set_column_widths(
        worksheet,
        {
            "A": 14,
            "B": 24,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 17,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 20,
        },
    )

    worksheet.freeze_panes = f"A{header_row + 1}"


def _create_fleet_sheet(
    workbook: Workbook,
    branding: ReportBranding,
    styles: Dict[str, Any],
    rows: List[Dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet("Fleet")
    worksheet.sheet_view.showGridLines = False

    subtitle = (
        f"{branding.company_name} | {branding.mine_name} | "
        "Fleet availability and utilization"
    )
    header_row = _style_title(
        worksheet,
        "Fleet Performance",
        subtitle,
        styles,
        5,
    )

    headers = [
        "Report Date",
        "Mine",
        "Availability",
        "Utilization",
        "Created At",
    ]

    _style_table_header(
        worksheet,
        header_row,
        headers,
        styles,
    )

    for row_index, record in enumerate(rows, start=header_row + 1):
        availability = _number(record.get("availability"))
        utilization = _number(record.get("utilization"))

        values = [
            record.get("report_date"),
            record.get("mine_name") or branding.mine_name,
            availability / 100,
            utilization / 100,
            record.get("created_at"),
        ]

        for column_index, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

        _style_data_row(
            worksheet,
            row_index,
            len(headers),
            styles,
        )

        _apply_date_format(worksheet.cell(row=row_index, column=1))
        _apply_percentage_format(
            worksheet.cell(row=row_index, column=3)
        )
        _apply_percentage_format(
            worksheet.cell(row=row_index, column=4)
        )
        _apply_datetime_format(
            worksheet.cell(row=row_index, column=5)
        )

    last_row = header_row + len(rows)
    _add_autofilter(
        worksheet,
        header_row,
        last_row,
        len(headers),
    )

    if last_row > header_row:
        for column_letter in ["C", "D"]:
            target_range = (
                f"{column_letter}{header_row + 1}:"
                f"{column_letter}{last_row}"
            )
            worksheet.conditional_formatting.add(
                target_range,
                CellIsRule(
                    operator="lessThan",
                    formula=["0.75"],
                    fill=styles["red_fill"],
                ),
            )
            worksheet.conditional_formatting.add(
                target_range,
                CellIsRule(
                    operator="between",
                    formula=["0.75", "0.849999"],
                    fill=styles["amber_fill"],
                ),
            )
            worksheet.conditional_formatting.add(
                target_range,
                CellIsRule(
                    operator="greaterThanOrEqual",
                    formula=["0.85"],
                    fill=styles["green_fill"],
                ),
            )

    _set_column_widths(
        worksheet,
        {
            "A": 14,
            "B": 26,
            "C": 18,
            "D": 18,
            "E": 20,
        },
    )

    worksheet.freeze_panes = f"A{header_row + 1}"


def _create_plant_sheet(
    workbook: Workbook,
    branding: ReportBranding,
    styles: Dict[str, Any],
    rows: List[Dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet("Plant")
    worksheet.sheet_view.showGridLines = False

    subtitle = (
        f"{branding.company_name} | {branding.mine_name} | "
        "Plant throughput and recovery"
    )
    header_row = _style_title(
        worksheet,
        "Plant Performance",
        subtitle,
        styles,
        8,
    )

    headers = [
        "Report Date",
        "Mine",
        "Throughput Plan",
        "Throughput Actual",
        "Variance",
        "Achievement",
        "Recovery",
        "Created At",
    ]

    _style_table_header(
        worksheet,
        header_row,
        headers,
        styles,
    )

    for row_index, record in enumerate(rows, start=header_row + 1):
        plan = _number(record.get("throughput_plan"))
        actual = _number(record.get("throughput_actual"))
        recovery = _number(record.get("recovery"))

        values = [
            record.get("report_date"),
            record.get("mine_name") or branding.mine_name,
            plan,
            actual,
            actual - plan,
            _safe_ratio(actual, plan),
            recovery / 100,
            record.get("created_at"),
        ]

        for column_index, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

        _style_data_row(
            worksheet,
            row_index,
            len(headers),
            styles,
        )

        _apply_date_format(worksheet.cell(row=row_index, column=1))

        for column_index in [3, 4, 5]:
            _apply_number_format(
                worksheet.cell(
                    row=row_index,
                    column=column_index,
                )
            )

        _apply_percentage_format(
            worksheet.cell(row=row_index, column=6)
        )
        _apply_percentage_format(
            worksheet.cell(row=row_index, column=7)
        )
        _apply_datetime_format(
            worksheet.cell(row=row_index, column=8)
        )

    last_row = header_row + len(rows)
    _add_autofilter(
        worksheet,
        header_row,
        last_row,
        len(headers),
    )

    if last_row > header_row:
        achievement_range = f"F{header_row + 1}:F{last_row}"
        worksheet.conditional_formatting.add(
            achievement_range,
            CellIsRule(
                operator="lessThan",
                formula=["0.90"],
                fill=styles["red_fill"],
            ),
        )
        worksheet.conditional_formatting.add(
            achievement_range,
            CellIsRule(
                operator="between",
                formula=["0.90", "0.999999"],
                fill=styles["amber_fill"],
            ),
        )
        worksheet.conditional_formatting.add(
            achievement_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["1.00"],
                fill=styles["green_fill"],
            ),
        )

    _set_column_widths(
        worksheet,
        {
            "A": 14,
            "B": 26,
            "C": 19,
            "D": 19,
            "E": 15,
            "F": 16,
            "G": 15,
            "H": 20,
        },
    )

    worksheet.freeze_panes = f"A{header_row + 1}"


def _create_safety_sheet(
    workbook: Workbook,
    branding: ReportBranding,
    styles: Dict[str, Any],
    rows: List[Dict[str, Any]],
) -> None:
    worksheet = workbook.create_sheet("Safety")
    worksheet.sheet_view.showGridLines = False

    subtitle = (
        f"{branding.company_name} | {branding.mine_name} | "
        "Safety performance and critical risk indicators"
    )
    header_row = _style_title(
        worksheet,
        "Safety Performance",
        subtitle,
        styles,
        7,
    )

    headers = [
        "Report Date",
        "Mine",
        "Incidents",
        "Near Misses",
        "Critical Risks",
        "Safety Score",
        "Created At",
    ]

    _style_table_header(
        worksheet,
        header_row,
        headers,
        styles,
    )

    for row_index, record in enumerate(rows, start=header_row + 1):
        values = [
            record.get("report_date"),
            record.get("mine_name") or branding.mine_name,
            int(_number(record.get("incidents"))),
            int(_number(record.get("near_misses"))),
            int(_number(record.get("critical_risks"))),
            _number(record.get("safety_score")) / 100,
            record.get("created_at"),
        ]

        for column_index, value in enumerate(values, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

        _style_data_row(
            worksheet,
            row_index,
            len(headers),
            styles,
        )

        _apply_date_format(worksheet.cell(row=row_index, column=1))
        _apply_percentage_format(
            worksheet.cell(row=row_index, column=6)
        )
        _apply_datetime_format(
            worksheet.cell(row=row_index, column=7)
        )

    last_row = header_row + len(rows)
    _add_autofilter(
        worksheet,
        header_row,
        last_row,
        len(headers),
    )

    if last_row > header_row:
        incidents_range = f"C{header_row + 1}:C{last_row}"
        risks_range = f"E{header_row + 1}:E{last_row}"
        score_range = f"F{header_row + 1}:F{last_row}"

        worksheet.conditional_formatting.add(
            incidents_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=styles["red_fill"],
            ),
        )
        worksheet.conditional_formatting.add(
            risks_range,
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=styles["amber_fill"],
            ),
        )
        worksheet.conditional_formatting.add(
            score_range,
            CellIsRule(
                operator="lessThan",
                formula=["0.90"],
                fill=styles["red_fill"],
            ),
        )
        worksheet.conditional_formatting.add(
            score_range,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=["0.95"],
                fill=styles["green_fill"],
            ),
        )

    _set_column_widths(
        worksheet,
        {
            "A": 14,
            "B": 26,
            "C": 14,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 20,
        },
    )

    worksheet.freeze_panes = f"A{header_row + 1}"


# -------------------------------------------------------------------
# KPI definitions worksheet
# -------------------------------------------------------------------

def _create_kpi_definitions_sheet(
    workbook: Workbook,
    branding: ReportBranding,
    styles: Dict[str, Any],
) -> None:
    worksheet = workbook.create_sheet("KPI Definitions")
    worksheet.sheet_view.showGridLines = False

    subtitle = (
        f"{branding.company_name} | {branding.mine_name} | "
        "Definitions used in this workbook"
    )
    header_row = _style_title(
        worksheet,
        "KPI Definitions",
        subtitle,
        styles,
        5,
    )

    headers = [
        "Category",
        "KPI",
        "Definition",
        "Calculation",
        "Interpretation",
    ]

    _style_table_header(
        worksheet,
        header_row,
        headers,
        styles,
    )

    definitions = [
        (
            "Production",
            "Ore Plan Achievement",
            "Actual ore movement compared with planned ore movement.",
            "Ore Actual ÷ Ore Plan",
            "100% or above indicates plan was achieved.",
        ),
        (
            "Production",
            "Waste Plan Achievement",
            "Actual waste movement compared with planned waste movement.",
            "Waste Actual ÷ Waste Plan",
            "100% or above indicates plan was achieved.",
        ),
        (
            "Fleet",
            "Availability",
            "Percentage of scheduled time that equipment is available.",
            "Available Time ÷ Scheduled Time",
            "Higher is generally better.",
        ),
        (
            "Fleet",
            "Utilization",
            "Percentage of available time that equipment is operating.",
            "Operating Time ÷ Available Time",
            "Higher is generally better.",
        ),
        (
            "Plant",
            "Throughput Achievement",
            "Actual processing throughput compared with plan.",
            "Throughput Actual ÷ Throughput Plan",
            "100% or above indicates plan was achieved.",
        ),
        (
            "Plant",
            "Recovery",
            "Percentage of valuable material recovered through processing.",
            "Recovered Value ÷ Feed Value",
            "Higher is generally better.",
        ),
        (
            "Safety",
            "Incidents",
            "Recorded safety incidents for the reporting date.",
            "Count",
            "Zero is preferred.",
        ),
        (
            "Safety",
            "Near Misses",
            "Reported events that could have caused harm or loss.",
            "Count",
            "Requires review and follow-up.",
        ),
        (
            "Safety",
            "Critical Risks",
            "Open or observed critical risk exposures.",
            "Count",
            "Zero unresolved critical exposures is preferred.",
        ),
        (
            "Safety",
            "Safety Score",
            "Composite safety performance indicator.",
            "Configured safety scoring methodology",
            "Higher is generally better.",
        ),
    ]

    for row_index, definition in enumerate(
        definitions,
        start=header_row + 1,
    ):
        for column_index, value in enumerate(definition, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

        _style_data_row(
            worksheet,
            row_index,
            len(headers),
            styles,
        )

        for column_index in range(1, len(headers) + 1):
            worksheet.cell(
                row=row_index,
                column=column_index,
            ).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    _set_column_widths(
        worksheet,
        {
            "A": 17,
            "B": 27,
            "C": 48,
            "D": 32,
            "E": 43,
        },
    )

    worksheet.freeze_panes = f"A{header_row + 1}"
    worksheet.auto_filter.ref = (
        f"A{header_row}:E{header_row + len(definitions)}"
    )


# -------------------------------------------------------------------
# Workbook metadata
# -------------------------------------------------------------------

def _set_workbook_properties(
    workbook: Workbook,
    branding: ReportBranding,
) -> None:
    workbook.properties.title = (
        f"{branding.company_name} Executive Operations Export"
    )
    workbook.properties.subject = (
        f"Operational KPI workbook for {branding.mine_name}"
    )
    workbook.properties.creator = "Mine Manager AI"
    workbook.properties.company = branding.company_name
    workbook.properties.description = (
        "Executive operations workbook generated from production, fleet, "
        "plant, and safety data."
    )
    workbook.properties.created = datetime.now()
    workbook.properties.modified = datetime.now()


# -------------------------------------------------------------------
# Public export function
# -------------------------------------------------------------------

def generate_executive_excel_export() -> BytesIO:
    """
    Generate the branded Mine Manager AI executive Excel workbook.

    Worksheets:
        1. Executive Summary
        2. Production
        3. Fleet
        4. Plant
        5. Safety
        6. KPI Definitions

    Branding is loaded from CompanySettings and MineSettings.
    Safe defaults are supplied by report_branding_service.
    """

    branding = get_report_branding()
    styles = _build_styles(branding)

    production_rows = _fetch_production_data()
    fleet_rows = _fetch_fleet_data()
    plant_rows = _fetch_plant_data()
    safety_rows = _fetch_safety_data()

    workbook = Workbook()
    _set_workbook_properties(workbook, branding)

    _create_executive_summary_sheet(
        workbook=workbook,
        branding=branding,
        styles=styles,
        production_rows=production_rows,
        fleet_rows=fleet_rows,
        plant_rows=plant_rows,
        safety_rows=safety_rows,
    )

    _create_production_sheet(
        workbook=workbook,
        branding=branding,
        styles=styles,
        rows=production_rows,
    )

    _create_fleet_sheet(
        workbook=workbook,
        branding=branding,
        styles=styles,
        rows=fleet_rows,
    )

    _create_plant_sheet(
        workbook=workbook,
        branding=branding,
        styles=styles,
        rows=plant_rows,
    )

    _create_safety_sheet(
        workbook=workbook,
        branding=branding,
        styles=styles,
        rows=safety_rows,
    )

    _create_kpi_definitions_sheet(
        workbook=workbook,
        branding=branding,
        styles=styles,
    )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer
