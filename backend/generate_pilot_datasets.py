from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path
import random

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent
PILOT_DATA_ROOT = PROJECT_ROOT / "pilot-package" / "data"
TEMPLATE_DIR = PILOT_DATA_ROOT / "templates"
DEMO_DIR = PILOT_DATA_ROOT / "demo"

SAMPLE_START_DATE = date(2026, 7, 1)
SAMPLE_DAYS = 30
RANDOM_SEED = 21021

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

SUBTITLE_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)


DATASET_DEFINITIONS = {
    "production": {
        "sheet_name": "Production",
        "template_filename": (
            "Mine_Manager_AI_Production_Template.xlsx"
        ),
        "sample_filename": (
            "Mine_Manager_AI_Production_30_Day_Sample.xlsx"
        ),
        "headers": [
            "report_date",
            "ore_plan",
            "ore_actual",
            "waste_plan",
            "waste_actual",
        ],
        "descriptions": [
            "Reporting date in YYYY-MM-DD format",
            "Planned ore movement",
            "Actual ore movement",
            "Planned waste movement",
            "Actual waste movement",
        ],
    },
    "fleet": {
        "sheet_name": "Fleet",
        "template_filename": (
            "Mine_Manager_AI_Fleet_Template.xlsx"
        ),
        "sample_filename": (
            "Mine_Manager_AI_Fleet_30_Day_Sample.xlsx"
        ),
        "headers": [
            "report_date",
            "truck_id",
            "availability",
            "utilization",
        ],
        "descriptions": [
            "Reporting date in YYYY-MM-DD format",
            "Truck identifier, for example TRK-101",
            "Fleet availability percentage from 0 to 100",
            "Fleet utilization percentage from 0 to 100",
        ],
    },
    "plant": {
        "sheet_name": "Plant",
        "template_filename": (
            "Mine_Manager_AI_Plant_Template.xlsx"
        ),
        "sample_filename": (
            "Mine_Manager_AI_Plant_30_Day_Sample.xlsx"
        ),
        "headers": [
            "report_date",
            "throughput_plan",
            "throughput_actual",
            "recovery",
        ],
        "descriptions": [
            "Reporting date in YYYY-MM-DD format",
            "Planned plant throughput",
            "Actual plant throughput",
            "Plant recovery percentage from 0 to 100",
        ],
    },
    "safety": {
        "sheet_name": "Safety",
        "template_filename": (
            "Mine_Manager_AI_Safety_Template.xlsx"
        ),
        "sample_filename": (
            "Mine_Manager_AI_Safety_30_Day_Sample.xlsx"
        ),
        "headers": [
            "report_date",
            "incidents",
            "near_misses",
            "critical_risks",
            "safety_score",
        ],
        "descriptions": [
            "Reporting date in YYYY-MM-DD format",
            "Number of safety incidents",
            "Number of near misses",
            "Number of critical-risk events",
            "Safety score percentage from 0 to 100",
        ],
    },
}


def ensure_directories() -> None:
    TEMPLATE_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    DEMO_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )


def style_header_row(
    worksheet,
    row_number: int,
    column_count: int,
) -> None:
    for column_index in range(
        1,
        column_count + 1,
    ):
        cell = worksheet.cell(
            row=row_number,
            column=column_index,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )


def set_column_widths(
    worksheet,
    headers: list[str],
) -> None:
    preferred_widths = {
        "report_date": 16,
        "truck_id": 15,
        "ore_plan": 15,
        "ore_actual": 15,
        "waste_plan": 15,
        "waste_actual": 15,
        "availability": 16,
        "utilization": 16,
        "throughput_plan": 20,
        "throughput_actual": 20,
        "recovery": 14,
        "incidents": 14,
        "near_misses": 16,
        "critical_risks": 17,
        "safety_score": 16,
    }

    for index, header in enumerate(
        headers,
        start=1,
    ):
        worksheet.column_dimensions[
            get_column_letter(index)
        ].width = preferred_widths.get(
            header,
            16,
        )


def add_metadata_sheet(
    workbook: Workbook,
    dataset_name: str,
    headers: list[str],
    descriptions: list[str],
    is_template: bool,
) -> None:
    worksheet = workbook.create_sheet(
        "Instructions"
    )

    worksheet["A1"] = (
        "Mine Manager AI Pilot Dataset"
    )
    worksheet["A1"].font = Font(
        bold=True,
        size=16,
    )

    worksheet["A3"] = "Dataset"
    worksheet["B3"] = (
        dataset_name.title()
    )

    worksheet["A4"] = "File type"
    worksheet["B4"] = (
        "Header-only template"
        if is_template
        else "Synthetic 30-day sample dataset"
    )

    worksheet["A5"] = (
        "Data classification"
    )
    worksheet["B5"] = (
        "Synthetic pilot data only"
    )

    worksheet["A6"] = "Date format"
    worksheet["B6"] = "YYYY-MM-DD"

    worksheet["A8"] = "Important"
    worksheet["A8"].font = Font(
        bold=True,
    )
    worksheet["A8"].fill = (
        SUBTITLE_FILL
    )

    worksheet["A9"] = (
        "Do not rename required columns. "
        "Do not add confidential customer data "
        "to sample files."
    )

    worksheet["A11"] = "Column"
    worksheet["B11"] = "Description"

    style_header_row(
        worksheet=worksheet,
        row_number=11,
        column_count=2,
    )

    for row_number, (
        header,
        description,
    ) in enumerate(
        zip(
            headers,
            descriptions,
        ),
        start=12,
    ):
        worksheet.cell(
            row=row_number,
            column=1,
            value=header,
        )

        worksheet.cell(
            row=row_number,
            column=2,
            value=description,
        )

    worksheet.column_dimensions[
        "A"
    ].width = 25

    worksheet.column_dimensions[
        "B"
    ].width = 65

    for row in range(
        1,
        worksheet.max_row + 1,
    ):
        worksheet.cell(
            row=row,
            column=1,
        ).alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )

        worksheet.cell(
            row=row,
            column=2,
        ).alignment = Alignment(
            vertical="top",
            wrap_text=True,
        )


def create_template(
    dataset_name: str,
    definition: dict,
) -> Path:
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = (
        definition["sheet_name"]
    )

    headers = definition["headers"]

    worksheet.append(headers)

    style_header_row(
        worksheet=worksheet,
        row_number=1,
        column_count=len(headers),
    )

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(headers))}"
        f"1"
    )

    set_column_widths(
        worksheet=worksheet,
        headers=headers,
    )

    add_metadata_sheet(
        workbook=workbook,
        dataset_name=dataset_name,
        headers=headers,
        descriptions=(
            definition["descriptions"]
        ),
        is_template=True,
    )

    output_path = (
        TEMPLATE_DIR
        / definition["template_filename"]
    )

    workbook.save(output_path)

    return output_path


def generate_production_rows(
    start_date: date,
    number_of_days: int,
) -> list[list]:
    rows = []

    for offset in range(
        number_of_days
    ):
        current_date = (
            start_date
            + timedelta(days=offset)
        )

        ore_plan = random.randint(
            8800,
            10200,
        )

        ore_actual = round(
            ore_plan
            * random.uniform(
                0.88,
                1.04,
            ),
            0,
        )

        waste_plan = random.randint(
            14500,
            17000,
        )

        waste_actual = round(
            waste_plan
            * random.uniform(
                0.90,
                1.06,
            ),
            0,
        )

        if offset in {8, 19}:
            ore_actual = round(
                ore_plan
                * random.uniform(
                    0.74,
                    0.82,
                ),
                0,
            )

        rows.append(
            [
                current_date,
                ore_plan,
                ore_actual,
                waste_plan,
                waste_actual,
            ]
        )

    return rows


def generate_fleet_rows(
    start_date: date,
    number_of_days: int,
) -> list[list]:
    rows = []

    truck_ids = [
        "TRK-101",
        "TRK-102",
        "TRK-103",
        "TRK-104",
        "TRK-105",
        "TRK-106",
    ]

    for offset in range(
        number_of_days
    ):
        current_date = (
            start_date
            + timedelta(days=offset)
        )

        truck_id = truck_ids[
            offset % len(truck_ids)
        ]

        availability = round(
            random.uniform(
                86.0,
                95.0,
            ),
            1,
        )

        utilization = round(
            random.uniform(
                76.0,
                91.0,
            ),
            1,
        )

        if offset in {8, 19}:
            availability = round(
                random.uniform(
                    78.0,
                    83.0,
                ),
                1,
            )

            utilization = round(
                random.uniform(
                    68.0,
                    75.0,
                ),
                1,
            )

        rows.append(
            [
                current_date,
                truck_id,
                availability,
                utilization,
            ]
        )

    return rows


def generate_plant_rows(
    start_date: date,
    number_of_days: int,
) -> list[list]:
    rows = []

    for offset in range(
        number_of_days
    ):
        current_date = (
            start_date
            + timedelta(days=offset)
        )

        throughput_plan = random.randint(
            30000,
            34000,
        )

        throughput_actual = round(
            throughput_plan
            * random.uniform(
                0.91,
                1.04,
            ),
            0,
        )

        recovery = round(
            random.uniform(
                87.0,
                92.5,
            ),
            1,
        )

        if offset in {8, 19}:
            throughput_actual = round(
                throughput_plan
                * random.uniform(
                    0.79,
                    0.86,
                ),
                0,
            )

            recovery = round(
                random.uniform(
                    82.0,
                    85.5,
                ),
                1,
            )

        rows.append(
            [
                current_date,
                throughput_plan,
                throughput_actual,
                recovery,
            ]
        )

    return rows


def generate_safety_rows(
    start_date: date,
    number_of_days: int,
) -> list[list]:
    rows = []

    for offset in range(
        number_of_days
    ):
        current_date = (
            start_date
            + timedelta(days=offset)
        )

        incidents = 0

        near_misses = random.choice(
            [0, 0, 0, 1, 1, 2]
        )

        critical_risks = random.choice(
            [0, 0, 0, 0, 1]
        )

        safety_score = round(
            random.uniform(
                94.0,
                99.5,
            ),
            1,
        )

        if offset == 8:
            incidents = 1
            near_misses = 2
            critical_risks = 1
            safety_score = 88.0

        if offset == 19:
            incidents = 0
            near_misses = 3
            critical_risks = 2
            safety_score = 90.5

        rows.append(
            [
                current_date,
                incidents,
                near_misses,
                critical_risks,
                safety_score,
            ]
        )

    return rows


def create_sample_workbook(
    dataset_name: str,
    definition: dict,
    rows: list[list],
) -> Path:
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = (
        definition["sheet_name"]
    )

    headers = definition["headers"]

    worksheet.append(headers)

    style_header_row(
        worksheet=worksheet,
        row_number=1,
        column_count=len(headers),
    )

    for row in rows:
        worksheet.append(row)

    for cell in worksheet["A"][1:]:
        cell.number_format = (
            "yyyy-mm-dd"
        )

    worksheet.freeze_panes = "A2"

    worksheet.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(headers))}"
        f"{worksheet.max_row}"
    )

    set_column_widths(
        worksheet=worksheet,
        headers=headers,
    )

    worksheet.row_dimensions[
        1
    ].height = 24

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    add_metadata_sheet(
        workbook=workbook,
        dataset_name=dataset_name,
        headers=headers,
        descriptions=(
            definition["descriptions"]
        ),
        is_template=False,
    )

    output_path = (
        DEMO_DIR
        / definition["sample_filename"]
    )

    workbook.save(output_path)

    return output_path


def validate_generated_workbook(
    file_path: Path,
    expected_headers: list[str],
    expected_data_rows: int,
) -> None:
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True,
    )

    worksheet = workbook[
        workbook.sheetnames[0]
    ]

    actual_headers = [
        cell.value
        for cell in next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
            )
        )
    ]

    if actual_headers != expected_headers:
        workbook.close()

        raise ValueError(
            f"Header mismatch in "
            f"{file_path.name}: "
            f"{actual_headers}"
        )

    actual_data_rows = max(
        worksheet.max_row - 1,
        0,
    )

    if (
        actual_data_rows
        != expected_data_rows
    ):
        workbook.close()

        raise ValueError(
            f"Unexpected row count in "
            f"{file_path.name}: "
            f"{actual_data_rows}"
        )

    workbook.close()


def validate_value_ranges(
    dataset_name: str,
    rows: list[list],
) -> None:
    if dataset_name == "production":
        for row in rows:
            numeric_values = row[1:]

            if any(
                float(value) < 0
                for value in numeric_values
            ):
                raise ValueError(
                    "Production values "
                    "cannot be negative."
                )

    elif dataset_name == "fleet":
        for row in rows:
            truck_id = row[1]
            availability = float(row[2])
            utilization = float(row[3])

            if not truck_id:
                raise ValueError(
                    "Fleet truck_id "
                    "cannot be empty."
                )

            if not (
                0 <= availability <= 100
            ):
                raise ValueError(
                    "Fleet availability "
                    "must be between 0 and 100."
                )

            if not (
                0 <= utilization <= 100
            ):
                raise ValueError(
                    "Fleet utilization "
                    "must be between 0 and 100."
                )

    elif dataset_name == "plant":
        for row in rows:
            throughput_plan = float(
                row[1]
            )
            throughput_actual = float(
                row[2]
            )
            recovery = float(
                row[3]
            )

            if (
                throughput_plan < 0
                or throughput_actual < 0
            ):
                raise ValueError(
                    "Plant throughput values "
                    "cannot be negative."
                )

            if not (
                0 <= recovery <= 100
            ):
                raise ValueError(
                    "Plant recovery must be "
                    "between 0 and 100."
                )

    elif dataset_name == "safety":
        for row in rows:
            incidents = int(row[1])
            near_misses = int(row[2])
            critical_risks = int(row[3])
            safety_score = float(row[4])

            if any(
                value < 0
                for value in [
                    incidents,
                    near_misses,
                    critical_risks,
                ]
            ):
                raise ValueError(
                    "Safety counts "
                    "cannot be negative."
                )

            if not (
                0 <= safety_score <= 100
            ):
                raise ValueError(
                    "Safety score must be "
                    "between 0 and 100."
                )


def main() -> None:
    random.seed(
        RANDOM_SEED
    )

    ensure_directories()

    generated_files: list[
        Path
    ] = []

    for (
        dataset_name,
        definition,
    ) in DATASET_DEFINITIONS.items():
        template_path = create_template(
            dataset_name=dataset_name,
            definition=definition,
        )

        validate_generated_workbook(
            file_path=template_path,
            expected_headers=(
                definition["headers"]
            ),
            expected_data_rows=0,
        )

        generated_files.append(
            template_path
        )

    sample_rows = {
        "production": (
            generate_production_rows(
                start_date=(
                    SAMPLE_START_DATE
                ),
                number_of_days=(
                    SAMPLE_DAYS
                ),
            )
        ),
        "fleet": (
            generate_fleet_rows(
                start_date=(
                    SAMPLE_START_DATE
                ),
                number_of_days=(
                    SAMPLE_DAYS
                ),
            )
        ),
        "plant": (
            generate_plant_rows(
                start_date=(
                    SAMPLE_START_DATE
                ),
                number_of_days=(
                    SAMPLE_DAYS
                ),
            )
        ),
        "safety": (
            generate_safety_rows(
                start_date=(
                    SAMPLE_START_DATE
                ),
                number_of_days=(
                    SAMPLE_DAYS
                ),
            )
        ),
    }

    for (
        dataset_name,
        rows,
    ) in sample_rows.items():
        validate_value_ranges(
            dataset_name=dataset_name,
            rows=rows,
        )

    for (
        dataset_name,
        definition,
    ) in DATASET_DEFINITIONS.items():
        sample_path = (
            create_sample_workbook(
                dataset_name=(
                    dataset_name
                ),
                definition=(
                    definition
                ),
                rows=(
                    sample_rows[
                        dataset_name
                    ]
                ),
            )
        )

        validate_generated_workbook(
            file_path=sample_path,
            expected_headers=(
                definition["headers"]
            ),
            expected_data_rows=(
                SAMPLE_DAYS
            ),
        )

        generated_files.append(
            sample_path
        )

    print(
        "Pilot datasets "
        "generated successfully."
    )
    print()

    for generated_file in (
        generated_files
    ):
        print(
            generated_file
        )


if __name__ == "__main__":
    main()